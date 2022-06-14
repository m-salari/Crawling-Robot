import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl


class Gmail:
    def __init__(self, sender_address, sender_password, text):
        self.sender_address = sender_address
        self.sender_pass = sender_password
        self.mail_content = text

    def Login(self):
        global session
        session = smtplib.SMTP('smtp.gmail.com', 587)  # use gmail with port
        session.starttls()  # enable security
        session.login(self.sender_address, self.sender_pass)  # login with mail_id and password
        print("login success!")

    def SendMail(self, receiver_address, subject):
        # global session
        message = MIMEMultipart()
        message['From'] = self.sender_address
        message['To'] = receiver_address
        message['Subject'] = subject  # The subject line
        message.attach(MIMEText(self.mail_content, 'plain'))

        text = message.as_string()
        session.sendmail(self.sender_address, receiver_address, text)
        print('message Sent to {0}'.format(receiver_address))

    def Exit(self):
        session.quit()

class Excel:
    def __init__(self, path):
        self.path = path
        self.wb_obj = openpyxl.load_workbook(self.path)
        self.sheet_obj = self.wb_obj.active

    def ReadEmail(self, number):
        flag = self.CheckingValidEmail(number)
        if flag == False:
            return False
        cell_obj = self.sheet_obj.cell(row=number, column=2)
        return cell_obj.value

    def Star(self, number):
        cell_obj = self.sheet_obj.cell(row=number, column=3)
        cell_obj.value = "*"
        self.wb_obj.save(self.path)

    def CheckingValidEmail(self, number):
        cell_obj = self.sheet_obj.cell(row=number, column=3)
        if cell_obj.value == "*":
            return False


def main(send_address, send_pass, content, subject, excel_name, count_of_email):
    pat = excel_name
    excel = Excel(pat)
    gmail = Gmail(send_address, send_pass, content)
    gmail.Login()

    for k in range(2, count_of_email + 2):
        mail = excel.ReadEmail(k)
        if mail:
            gmail.SendMail(mail, subject)
            excel.Star(k)
        time.sleep(1)
    gmail.Exit()


if __name__ == '__main__':
    send_address = 'email'
    send_pass = 'password'
    main(send_address, send_pass, "hello", "this is a test", "sample.xlsx", 2)
